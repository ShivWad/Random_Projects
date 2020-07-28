import requests
from bs4 import BeautifulSoup as bs
from collections import defaultdict
import pandas as pd
import openpyxl
alphabets = [chr(x) for x in range(ord('A'), ord('Z') + 1)]
alphabets.append('0')
Disease=defaultdict(list)
link = 'https://www.mayoclinic.org/diseases-conditions/index?letter='
names=[]
for i in alphabets:
    New_link = link +i
    req = requests.get(New_link)
    info = bs(req.text , 'html.parser')
    disease = info.find(class_ = 'index content-within')
    try:
        for j in disease.find_all('li'):

            name_link = 'https://www.mayoclinic.org' + str(j.a['href'])

            req = requests.get(name_link, allow_redirects=False ,headers={'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_9_2) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/34.0.1847.131 Safari/537.36'})
            print(req.status_code)
            print(req.url)
            data = bs(req.text,'html.parser')
            n = data.find('h1')
            names.append(n.text)
    except AttributeError as ae:
        print('Caught error')
        continue

for i in alphabets:
    for j in names:
        if i==j[0][0]:
            Disease[i].append(j)
##Converting into xlsx
workbook = openpyxl.Workbook()
sheet = workbook.active
row = 1
for key,values in Disease.items():
    # Put the key in the first column for each key in the dictionary
    sheet.cell(row=row, column=1, value=key)
    column = 2
    for element in values:
        # Put the element in each adjacent column for each element in the tuple
        sheet.cell(row=row, column=column, value=element)
        column += 1
    row += 1
workbook.save(filename='Disease_info.xlsx')
